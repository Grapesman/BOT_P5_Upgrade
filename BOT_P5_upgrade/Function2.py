from openpyxl import load_workbook
import requests, os, datetime
from dotenv import load_dotenv
load_dotenv()

def function2():
    # # СПИСОК ДЛЯ РЕАЛИЗАЦИИ 2 КОМАНДЫ
    a = []
    list_date_states = []  # Список для парсинга дат публикации статей с листа "Каталог статей"
    list_name_states = []  # Список для парсинга названий статей с листа "Каталог статей"
    list_make_states = []  # Список для парсинга дат создания статей с листа "Каталог статей"
    dict_date_name_state = {}  # словарь для объединения дат публикации и названий статей с листа "Каталог статей"
    dict_make_name_state = {}  # словарь для объединения дат создания и названий статей с листа "Каталог статей"
    date_2_list = []  # Список для сохранения дат статей, 6 и менее месяца от текущий с листа "Каталог статей"
    search_state = []  # Список статей для вывода в бот
    make_state = []  # Список дат создания по названию статей для вывода в бот
    global check_state_in_dict
    global date_check_make_in_dict
    global date_state_3m
    # Столбцы с названиями
    # B - "Название"
    # D - "Даты публикации"
    # F - "Тезиса"
    # H - "Авторы"
    # J - "Ключевые слова"

    # ----------------------------------------------------------------------------------------------------------------------
    # --------------------------------------------------ФУНКЦИИ 2 КОМАНДЫ---------------------------------------------------

    # РАЗДЕЛ ДЛЯ 2 КОМАНДЫ: ВЫВОД СТАТЕЙ НА БЛИЖАЙШИЕ 3 МЕСЯЦА

    # Функция определения количества заполненных строк. Определяем по столбцу "A" с нумерацией для листа "Каталог статей"
    def how_much_string(file_path):
        i = 1
        book = load_workbook(filename = file_path)
        sheet = book['Каталог статей']
        while sheet['A' + str(i)].value is not None:
            a.append(i)
            i+=1
        return
    # функции для парсинга даты публикации с листа "Каталог статей"
    def date_check(file_path):
        book = load_workbook(filename=file_path)
        sheet = book['Каталог статей']
        for i in range(2, len(a) + 1):
            if sheet['D' + str(i)].value is not None and sheet['C' + str(i)].value is not None and sheet[
                'B' + str(i)].value is not None:
                list_date_state = sheet['D' + str(i)].value
                list_date_states.append(list_date_state)
        return list_date_states

    # функции для парсинга названия статьи с листа "Каталог статей"
    def name_check(file_path):
        book = load_workbook(filename=file_path)
        sheet = book['Каталог статей']
        for i in range(2, len(a) + 1):
            if sheet['D' + str(i)].value is not None and sheet['C' + str(i)].value is not None and sheet[
                'B' + str(i)].value is not None:
                list_name_state = sheet['B' + str(i)].value
                list_name_states.append(list_name_state)
        return list_name_states

    # функции для парсинга даты создания с листа "Каталог статей"
    def make_check(file_path):
        book = load_workbook(filename=file_path)
        sheet = book['Каталог статей']
        for i in range(2, len(a) + 1):
            if sheet['D' + str(i)].value is not None and sheet['C' + str(i)].value is not None and sheet[
                'B' + str(i)].value is not None:
                list_make_state = sheet['C' + str(i)].value
                list_make_states.append(list_make_state)
        return list_make_states

    # Функция для определения сегодняшней даты
    def get_todays_date():
        today = datetime.datetime.today()
        return today.strftime("%Y-%m-%d")

    # Функция определения разницы между сегодняшней датой и даты из списка
    def compare_dates():
        date1 = datetime.datetime.strptime(x, "%Y-%m-%d")
        for i in range(len(DATA)):
            date2 = DATA[i]
            if 0 < (date2 - date1).days <= 180:
                date_2_list.append(date2)
        return date_2_list

    # Проверяем, есть ли статьи по данным датам публикации
    def state_function(checking):
        for i in range(0, len(checking)):
            if checking[i] in dict_date_name_state:
                dic = dict_date_name_state[checking[i]]
                search_state.append(dic)
        return search_state

    # Запускаем функцию проверки наличия даты создания в словаре по названию статей
    def make_function(state_3m):
        for i in range(0, len(state_3m)):
            if state_3m[i] in dict_make_name_state:
                dic = dict_make_name_state[state_3m[i]]
                make_state.append(dic)
        return make_state

    # Убираем повторяющиеся названия статей
    def remove_duplicates_3m(search_state):
        # Используем множество для хранения уникальных слов
        unique_state_3m = list(set(search_state))
        # Преобразуем обратно в список
        return unique_state_3m

# ----------------------------------------------------------------------------------------------------------------------
# ------------------------------------------ВЫПОЛНЕНИЕ ФУНКЦИЙ ДЛЯ 2 КОМАНДЫ--------------------------------------------

    # Считаем количество строк в файле по листу "Каталог статей"
    stroki = how_much_string(os.getenv('SAVE_PATH'))
    # Определяем текущую дату
    x = get_todays_date()
    # Сохраняем список дат публикации статей
    DATA = date_check(os.getenv('SAVE_PATH'))
    # Сохраняем названия статей
    STATE = name_check(os.getenv('SAVE_PATH'))
    # Сохраняем список дат создания статей
    MAKE = make_check(os.getenv('SAVE_PATH'))

    # Объединяем в словарь список дат публикации статей и их названия
    dict_date_name_state = dict(zip(DATA, STATE))
    # Объединяем в словарь список дат создания статей и их названия
    dict_make_name_state = dict(zip(STATE, MAKE))

    # Определяем даты публикации, относительно текущей, в диапазоне до 180 дней и сохраняем в виде списка date_2_list
    checking = compare_dates()

    # Убираем повторяющиеся даты из списка
    state_3m = remove_duplicates_3m(date_2_list)

    # Сохраняем список статей по данным датам
    check_state_in_dict = state_function(state_3m)

    # Находим даты создания статьи по её названию
    check_make_in_dict = make_function(check_state_in_dict)

    # Переводим списки дат создания и публикации из формата date в формат datetime
    date_check_make_in_dict = [dt.date() for dt in check_make_in_dict]
    date_state_3m = [dt.date() for dt in state_3m]