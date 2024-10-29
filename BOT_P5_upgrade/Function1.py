from openpyxl import load_workbook
import os
from dotenv import load_dotenv
load_dotenv()
def function1():
    # СПИСОК ДЛЯ РЕАЛИЗАЦИИ 1 КОМАНДЫ
    # Вводим списки
    a = [] # Запоминает количество заполненных в таблице строк на листе "Каталог статей"
    b = [] # Запоминает количество заполненных в таблице строк на листе "Community"
    list_names_input = [] # Список для сохранения значений фамилий с листа "Community"
    list_tegs_input = [] # Список для сохранения значений тегов с листа "Community"

    dict_name_tegs = {} # Словарь для сохранения значений фамилий и тегов с листа "Community"

    # СПИСОК ДЛЯ РЕАЛИЗАЦИИ 1-ГО УСЛОВИЯ (Есть автор, но нет названия)
    list_FIOs_without_name=[] # Список для сохранения авторов по первому условию
    list_Fs_without_name=[] # Список для разделения предыдущего по пробелу
    list_Familys_without_name = [] # Список оставляет только первое слово из предыдущего списка

    # СПИСОК ДЛЯ РЕАЛИЗАЦИИ 2-ГО УСЛОВИЯ (Есть автор, но нет даты публикации)
    list_FIOs_without_date=[] # Список для сохранения авторов по второму условию
    list_Fs_without_date=[] # Список для разделения предыдущего по пробелу
    list_Familys_without_date = [] # Список оставляет только первое слово из предыдущего списка

    # СПИСОК ДЛЯ РЕАЛИЗАЦИИ 3-ГО УСЛОВИЯ (Есть автор, но нет тезиса)
    list_FIOs_without_thesis=[] # Список для сохранения авторов по третьему условию
    list_Fs_without_thesis=[] # Список для разделения предыдущего по пробелу
    list_Familys_without_thesis = [] # Список оставляет только первое слово из предыдущего списка

    # СПИСОК ДЛЯ РЕАЛИЗАЦИИ 4-ГО УСЛОВИЯ (Есть автор, но нет Ключевых слов)
    list_FIOs_without_key=[] # Список для сохранения авторов по четвертому условию
    list_Fs_without_key=[] # Список для разделения предыдущего по пробелу
    list_Familys_without_key = [] # Список оставляет только первое слово из предыдущего списка

    # СПИСОК ДЛЯ РЕАЛИЗАЦИИ 5-ГО УСЛОВИЯ (Есть название, но нет Автора)
    list_state_names = [] # Список для сохранения авторов по пятому условию

    # Список для сохранения фамилий с 1-4 условий в один список
    list_result_Familys = []
    # Список для сохранения тегов по результатам списка list_result_Familys
    search_teg = []
    global all_teg
    global names_state
    # ----------------------------------------------------------------------------------------------------------------------
    # #Столбцы с названиями
    #         # B - "Название"
    #         # D - "Даты публикации"
    #         # F - "Тезиса"
    #         # H - "Авторы"
    #         # J - "Ключевые слова"
    # ----------------------------------------------------------------------------------------------------------------------
    # ----------------------------------------------------ФУНКЦИИ 1 КОМАНДЫ-------------------------------------------------

    # РАЗДЕЛ ДЛЯ 1 КОМАНДЫ: ВЫВОД АВТОРОВ И НАЗВАНИЙ СТАТЕЙ ПРИ НЕЗАПОЛНЕННОЙ ТАБЛИЦЕ
    # Определяем количество заполненных строк в файле по фамилиям для листа "Community"
    def how_much_string_community(file_path):
        i = 3
        book = load_workbook(filename = file_path)
        sheet = book['Community']
        while sheet['B' + str(i)].value is not None:
            b.append(i)
            i+=1
        return

    # Функция создания словаря при парсинге Фамилий и тегов
    def make_dict(file_path):
        global dict_name_tegs
        book = load_workbook(filename = file_path)
        sheet = book['Community']
        for i in range(3, len(b)+3):
            if sheet['B' + str(i)].value is not None and sheet['O' + str(i)].value is not None:
                list_name_input = sheet['B' + str(i)].value
                list_names_input.append(list_name_input)
                list_teg_input = sheet['O' + str(i)].value
                list_tegs_input.append(list_teg_input)
                dict_name_tegs = dict(zip(list_names_input, list_tegs_input))
        return dict_name_tegs

    # Функция определения количества заполненных строк. Определяем по столбцу "A" с нумерацией для листа "Каталог статей"
    def how_much_string(file_path):
        i = 1
        book = load_workbook(filename = file_path)
        sheet = book['Каталог статей']
        while sheet['A' + str(i)].value is not None:
            a.append(i)
            i+=1
        return

    # 1-Е УСЛОВИЕ. ЕСТЬ АВТОР, НО НЕТ НАЗВАНИЯ
    # Функция находит строки по данному условию и сохраняет список Авторов
    def f_FIOs_without_name(file_path):
        book = load_workbook(filename = file_path)
        sheet = book['Каталог статей']
        for i in range (2, len(a)+1):
            #Записал так, так как пустых строк нет, а есть непонятные символы - впоследствии исправим только на пустые строки
            if sheet['H' + str(i)].value is not None and sheet['B' + str(i)].value is None or sheet['B' + str(i)].value == 1 or sheet['B' + str(i)].value == 2 or sheet['B' + str(i)].value == 3 or sheet['B' + str(i)].value == '???':
                list_FIO_without_name = sheet['H' + str(i)].value
                list_FIOs_without_name.append(list_FIO_without_name)
        return list_FIOs_without_name
    # Функция в списке Авторов разделяет по пробелу элементы списка
    def f_Fs_without_name(list_FIOs_without_name):
        for i in range(0,len(list_FIOs_without_name)):
            list_F_without_name = (list_FIOs_without_name[i]).split()
            list_Fs_without_name.append(list_F_without_name)
        return list_Fs_without_name
    # Функция в списке Авторов оставляет только первое слово из элемента списка = фамилию
    def f_Familys_without_name(list_Fs_without_name):
        for i in range(0, len(list_Fs_without_name)):
            Family_without_name = (list_Fs_without_name[i])
            list_Family_without_name = Family_without_name[0]
            list_Familys_without_name.append(list_Family_without_name)
        return list_Familys_without_name

    # 2-Е УСЛОВИЕ. ЕСТЬ АВТОР, НО НЕТ ДАТЫ ПУБЛИКАЦИИ
    # Функция находит строки по данному условию и сохраняет список Авторов
    def f_FIOs_without_date(file_path):
        book = load_workbook(filename = file_path)
        sheet = book['Каталог статей']
        for i in range (2, len(a)+1):
            #Записал так, так как пустых строк нет, а есть непонятные символы - впоследствии исправим
            if sheet['H' + str(i)].value is not None and sheet['D' + str(i)].value is None:
                list_FIO_without_date = sheet['H' + str(i)].value
                list_FIOs_without_date.append(list_FIO_without_date)
        return list_FIOs_without_date
    # Функция в списке Авторов разделяет по пробелу элементы списка
    def f_Fs_without_date(list_FIOs_without_date):
        for i in range(0,len(list_FIOs_without_date)):
            list_F_without_date = (list_FIOs_without_date[i]).split()
            list_Fs_without_date.append(list_F_without_date)
        return list_Fs_without_date
    # Функция в списке Авторов оставляет только первое слово из элемента списка = фамилию
    def f_Familys_without_date(list_Fs_without_date):
        for i in range(0, len(list_Fs_without_date)):
            Family_without_date = (list_Fs_without_date[i])
            list_Family_without_date = Family_without_date[0]
            list_Familys_without_date.append(list_Family_without_date)
        return list_Familys_without_date

    # 3-Е УСЛОВИЕ. ЕСТЬ АВТОР, НО НЕТ ТЕЗИСА
    # Функция находит строки по данному условию и сохраняет список Авторов
    def f_FIOs_without_thesis(file_path):
        book = load_workbook(filename = file_path)
        sheet = book['Каталог статей']
        for i in range (2, len(a)+1):
            #Записал так, так как пустых строк нет, а есть непонятные символы - впоследствии исправим
            if sheet['H' + str(i)].value is not None and sheet['F' + str(i)].value is None:
                list_FIO_without_thesis = sheet['H' + str(i)].value
                list_FIOs_without_thesis.append(list_FIO_without_thesis)
        return list_FIOs_without_thesis
    # Функция в списке Авторов разделяет по пробелу элементы списка
    def f_Fs_without_thesis(list_FIOs_without_thesis):
        for i in range(0,len(list_FIOs_without_thesis)):
            list_F_without_thesis = (list_FIOs_without_thesis[i]).split()
            list_Fs_without_thesis.append(list_F_without_thesis)
        return list_Fs_without_thesis
    # Функция в списке Авторов оставляет только первое слово из элемента списка = фамилию
    def f_Familys_without_thesis(list_Fs_without_thesis):
        for i in range(0, len(list_Fs_without_thesis)):
            Family_without_thesis = (list_Fs_without_thesis[i])
            list_Family_without_thesis = Family_without_thesis[0]
            list_Familys_without_thesis.append(list_Family_without_thesis)
        return list_Familys_without_thesis

    # 4-Е УСЛОВИЕ. ЕСТЬ АВТОР, НО НЕТ КЛЮЧЕВЫХ СЛОВ
    # Функция находит строки по данному условию и сохраняет список Авторов
    def f_FIOs_without_key(file_path):
        book = load_workbook(filename = file_path)
        sheet = book['Каталог статей']
        for i in range (2, len(a)+1):
            #Записал так, так как пустых строк нет, а есть непонятные символы - впоследствии исправим
            if sheet['H' + str(i)].value is not None and sheet['J' + str(i)].value is None:
                list_FIO_without_key = sheet['H' + str(i)].value
                list_FIOs_without_key.append(list_FIO_without_key)
        return list_FIOs_without_key
    # Функция в списке Авторов разделяет по пробелу элементы списка
    def f_Fs_without_key(list_FIOs_without_key):
        for i in range(0,len(list_FIOs_without_key)):
            list_F_without_key = (list_FIOs_without_key[i]).split()
            list_Fs_without_key.append(list_F_without_key)
        return list_Fs_without_key
    # Функция в списке Авторов оставляет только первое слово из элемента списка = фамилию
    def f_Familys_without_key(list_Fs_without_key):
        for i in range(0, len(list_Fs_without_key)):
            Family_without_key = (list_Fs_without_key[i])
            list_Family_without_key = Family_without_key[0]
            list_Familys_without_key.append(list_Family_without_key)
        return list_Familys_without_key

    # Функция объединяет все получившиеся списки 1 УСЛОВИЯ с фамилиями в один
    def f_result_Familys (list_Familys_without_name, list_Familys_without_date, list_Familys_without_thesis, list_Familys_without_key):
        list_result_Familys = (list_Familys_without_name+ list_Familys_without_date + list_Familys_without_thesis +list_Familys_without_key)
        return list_result_Familys

    #Функция для удаления одинаковых фамилий в списке
    def remove_duplicates(list_result_Familys):
        # Используем множество для хранения уникальных слов
        unique_words = list(set(list_result_Familys))
        # Преобразуем обратно в список
        return unique_words

    # Функция нахождения тегов ТГ по фамилиям списка
    def teg_function(unique_words):
            for i in range (0, len(unique_words)):
                if unique_words[i] in data_dict:
                    teg = data_dict[unique_words[i]]
                    search_teg.append(teg)
            return search_teg

    # ----------------------------------------------------------------------------------------------------------------------
    # 5-Е УСЛОВИЕ. АВТОРА НЕТ, НО ЕСТЬ НАЗВАНИЕ
    def f_state_name(file_path):
        book = load_workbook(filename = file_path)
        sheet = book['Каталог статей']
        for i in range(2, len(a)+1):
            if sheet['H' + str(i)].value is None and sheet['B' + str(i)].value is not None:
                list_state_name = sheet['B' + str(i)].value
                list_state_names.append(list_state_name)
        return list_state_names

    # ----------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------ВЫПОЛНЕНИЕ ФУНКЦИЙ ДЛЯ 1 КОМАНДЫ--------------------------------------------

    # Считаем количество строк в файле по листу "Каталог статей"
    stroki = how_much_string(os.getenv('SAVE_PATH'))
    # Считаем количество строк в файле по листу "Community"
    stroki_community = how_much_string_community(os.getenv('SAVE_PATH'))

    # Формируем словарь с фамилиями и тегами с листа "Community"
    data_dict = make_dict(os.getenv('SAVE_PATH'))

    # Проверяем поочередно условия для листа "Каталог статей"
    # Проверка первого условия
    FIOs_without_name = f_FIOs_without_name(os.getenv('SAVE_PATH'))
    Fs_without_name = f_Fs_without_name(list_FIOs_without_name)
    Familys_without_name = f_Familys_without_name(list_Fs_without_name)

    # Проверка второго условия
    FIOs_without_date = f_FIOs_without_date(os.getenv('SAVE_PATH'))
    Fs_without_date = f_Fs_without_date(list_FIOs_without_date)
    Familys_without_date = f_Familys_without_date(list_Fs_without_date)

    # Проверка третьего условия
    FIOs_without_thesis = f_FIOs_without_thesis(os.getenv('SAVE_PATH'))
    Fs_without_thesis = f_Fs_without_thesis(list_FIOs_without_thesis)
    Familys_without_thesis = f_Familys_without_thesis(list_Fs_without_thesis)

    # Проверка четвертого условия
    FIOs_without_key = f_FIOs_without_key(os.getenv('SAVE_PATH'))
    Fs_without_key = f_Fs_without_key(list_FIOs_without_key)
    Familys_without_key = f_Familys_without_key(list_Fs_without_key)

    # Проверка пятого условия
    # Находим статьи, с заполненным названием, но без автора
    names_state = f_state_name(os.getenv('SAVE_PATH')) #<--------- это выводим в бот в команде ["status"]

    # Объединяем все полученные списки (1-4 условий) с фамилиями в один список
    result = f_result_Familys (list_Familys_without_name, list_Familys_without_date, list_Familys_without_thesis, list_Familys_without_key)
    # Исключаем повторы фамилий в списке
    without_duplication = remove_duplicates(result)
    # По получившимся фамилиям находим теги из словаря
    all_teg = teg_function(without_duplication) #<--------- это выводим в бот в команде ["status"]